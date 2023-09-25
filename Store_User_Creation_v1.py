import os
import re
import shutil
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill


def create_folder(folder_name: str) -> None:
	"""
	:rtype: Creates the folder of the respective company.
	"""
	try:
		os.makedirs(os.path.join(dir_path, folder_name), exist_ok=True)
	except FileExistsError:
		pass


def create_excel_sheet(excel_book: Workbook) -> Workbook:
	ws = excel_book.create_sheet(brand)
	ws['B1'] = f"{incident} provided access to users"
	ws['C1'] = incident
	ws['B2'] = "Roles"
	ws['C2'] = "GRC Requests"
	ws.merge_cells("A1:A2")
	ws['A1'] = "No."
	
	print(f"[INFO] Created {brand} sheet in the Excel Workbook for {user_id}")
	return excel_book


def cell_style(sheet: Worksheet) -> None:
	color = 'FFFF00'  # Yellow color
	
	for row in range(1, 3):
		for cell in sheet[row]:
			cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
	
	for row in sheet.iter_rows():
		for cell in row:
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.font = Font(name='Cascadia Code')
	
	for column in sheet.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = (max_length + 2) * 1.2
		sheet.column_dimensions[column[0].column_letter].width = adjusted_width
	
	sheet.freeze_panes = 'A3'
	sheet.sheet_view.zoomScale = 180


def role_specific_files(excel_data: str, workbook: Workbook) -> None:
	"""
	:rtype: Displays the Company, Report file and the number of roles within the company.
	"""
	if brand in ['VANS', 'TNF', 'TBL']:
		create_folder(brand)
	
	workbook = create_excel_sheet(workbook)
	sheet = workbook[brand]
	
	df = pd.read_csv(dir_path + excel_data)
	df = df[['User Login', 'Email', 'Role']]
	df = df.sort_values(by=['Role'], ascending=True, ignore_index=True)
	roles = df['Role'].unique()
	total_roles = roles.size
	
	for row, role in enumerate(roles, start=3):
		sheet.cell(row=row, column=1).value = total_roles - row + 3
		sheet.cell(row=row, column=2).value = role
	
	cell_style(sheet)
	
	template_df = pd.read_csv('Create_User_Template.csv')
	
	for role in roles:
		role_df = df[df['Role'] == role]
		temp = template_df.copy()
		temp['USERID'] = role_df['User Login']
		temp['MANAGER'] = user_id
		temp['EMAIL'] = role_df['Email']
		temp['SNC_NAME'] = 'p:CN=#!#USERID#!#'
		temp['UNSEC_SNC'] = 'Y'
		temp.to_csv(os.path.join(dir_path, brand, role + '.csv'), index=False)
	
	print(f"[INFO] Found {total_roles} {brand} Business roles.")


def find_brand(report_name: str) -> str:
	"""
	:rtype: Finding the Brand name from the file using Regular Expression.
	"""
	return re.search("[A-Z]{3,4}", report_name).group()


if __name__ == "__main__":
	dir_path = 'C:\\Users\\Dipaditya\\Downloads\\'

	wb = Workbook()

	reports = [key for key in os.listdir(dir_path) if key.startswith('report')]
	
	if reports:
		incident = input("Enter the Incident ID : ")
		excel_file = dir_path + incident + '_Store_User_Creation.xlsx'
		# user_id = input("Enter the PG1 Firefighter ID : ")
		
		print("=" * 70)
		print(f"All the Reports of {incident}:")
		print("-" * 70)
		for idx, file in enumerate(reports, start=1):
			print(f"{idx}.  {file}")
		
		print("=" * 70)
		print("Generating Log....")
		print("-" * 70)
		
		users = ['FF_SEC_2', 'FF_SEC_3']
		
		for user_id in users:
			create_folder(user_id)
			for report in reports:
				brand = find_brand(report)
				role_specific_files(report, wb)
				shutil.move(src=dir_path + brand, dst=dir_path + user_id)
				shutil.make_archive(base_name=dir_path + user_id, format='zip', root_dir=dir_path + user_id)
			
			shutil.rmtree(dir_path + user_id)
	
			del wb[wb.sheetnames[0]]
			wb.save(excel_file)
		print(f"[INFO] Successfully created {excel_file} and {len(reports)} zip files containing Role-Specific files.")
	else:
		print(f"[INFO] No report files present in {dir_path}")
