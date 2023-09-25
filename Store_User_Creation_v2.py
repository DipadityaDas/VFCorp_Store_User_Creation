import os
import re
import shutil
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_folder(folder_name: str) -> None:
	"""
	:rtype: Creates the folder of the respective company.
	"""
	try:
		os.makedirs(os.path.join(folder_path, folder_name), exist_ok=True)
	except FileExistsError:
		pass


def create_excel_sheet():
	ws = wb.create_sheet(brand)
	ws['B1'] = f"{incident} provided access to users"
	ws['C1'] = incident
	ws['B2'] = "Roles"
	ws['C2'] = "GRC Requests"
	ws.merge_cells("A1:A2")
	ws['A1'] = "No."
	
	df = unique_roles()
	roles = df['Role'].unique()
	total_roles = roles.size
	
	for row, role in enumerate(roles, start=3):
		ws.cell(row=row, column=1).value = total_roles - row + 3
		ws.cell(row=row, column=2).value = role
	
	color = 'FFFF00'  # Yellow color
	
	for row in range(1, 3):
		for cell in ws[row]:
			cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
	
	for row in ws.iter_rows():
		for cell in row:
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.font = Font(name='Cascadia Code')
	
	for column in ws.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = (max_length + 2) * 1.2
		ws.column_dimensions[column[0].column_letter].width = adjusted_width
	
	ws.freeze_panes = 'A3'
	ws.sheet_view.zoomScale = 180
	
	print(f"[INFO] Created {brand} sheet in the Excel Workbook")


def unique_roles():
	df = pd.read_csv(folder_path + report)
	df = df[necessary_details]
	df = df.sort_values(by=['Role'], ascending=True, ignore_index=True)
	return df


def roles_specific_files():
	"""
	:rtype: Displays the Company, Report file and the number of roles within the company.
	"""
	if brand in ['VANS', 'TNF', 'TBL']:
		create_folder(brand)
	
	df = unique_roles()
	roles = df['Role'].unique()
	total_roles = roles.size
	
	template_df = pd.read_csv('Create_User_Template.csv')
	
	for role in roles:
		role_df = df[df['Role'] == role]
		temp = template_df.copy()
		temp['USERID'] = role_df['User Login']
		temp['MANAGER'] = user_id
		temp['EMAIL'] = role_df['Email']
		temp['FNAME'] = role_df['Prefered First Name']
		temp['LNAME'] = role_df['Prefered Last Name']
		temp['SNC_NAME'] = 'p:CN=#!#USERID#!#'
		temp['UNSEC_SNC'] = 'Y'
		temp.to_csv(os.path.join(folder_path, brand, role + '.csv'), index=False)
	
	print(f"[INFO] Created {total_roles:3d} {brand:>4s} Business roles for {user_id}.")


def display_reports():
	print("=" * 70)
	print(f"All the Reports of {incident}:")
	print("-" * 70)
	for idx, file in enumerate(reports, start=1):
		print(f"{idx}.  {file}")
		
		
def generate_logs():
	print("=" * 70)
	print("Generating Logs....")
	print("-" * 70)


if __name__ == "__main__":
	folder_path = 'C:\\Users\\Dipaditya\\Downloads\\'
	reports = [key for key in os.listdir(folder_path) if key.startswith('report')]
	
	if reports:
		incident = input("Enter the Incident ID : ")
		excel_file = folder_path + incident + '_Store_User_Creation.xlsx'
		necessary_details = ['User Login', 'Email', 'Role', 'Prefered First Name', 'Prefered Last Name']
		ff_ids = ['FF_SEC_1', 'FF_SEC_2', 'FF_SEC_3']
		
		display_reports()
		generate_logs()
		
		wb = Workbook()
		c = 1
		for user_id in ff_ids:
			create_folder(user_id)
			
			for report in reports:
				brand = re.search("[A-Z]{3,4}", report).group()
				
				if c == 1:
					create_excel_sheet()
					
				roles_specific_files()
				
				shutil.move(src=folder_path + brand, dst=folder_path + user_id)
				shutil.make_archive(base_name=folder_path + user_id, format='zip', root_dir=folder_path + user_id)
			
			shutil.rmtree(folder_path + user_id)
			c += 1
		
		del wb[wb.sheetnames[0]]
		wb.save(excel_file)
		print(f"[INFO] Successfully created {excel_file}.")
		print(f"[INFO] Total Number of Role-Specific Zip file(s) : {len(ff_ids)}")
	
	else:
		print(f'[INFO] No reports present in the folder {folder_path}')
